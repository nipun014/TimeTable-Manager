"""Spreadsheet in, spreadsheet out. openpyxl was already a dependency."""
import csv
import io

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_BYTES = 5 * 1024 * 1024
MAX_ROWS = 5000

DAY_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

TEMPLATE = {
    "Classes": (
        ["class"],
        [["S3-CSE-A"], ["S3-CSE-B"]],
    ),
    "Subjects": (
        ["code", "name", "hours_per_week", "room_type", "block_size", "is_heavy"],
        [
            ["MAT203", "Discrete Mathematics", 4, "standard", 1, "yes"],
            ["CSL201", "Data Structures Lab", 3, "computer_lab", 3, "no"],
        ],
    ),
    "Teachers": (
        ["id", "name", "department", "can_teach", "unavailable"],
        [
            ["MAT01", "Dr. Anitha R", "Mathematics", "MAT203; MAT201", "Mon P1; Mon P2"],
            ["CSE04", "Prof. Vinod K", "Computer Science", "CSL201", ""],
        ],
    ),
    "Rooms": (
        ["id", "type", "capacity"],
        [["CR301", "standard", 60], ["CCL1", "computer_lab", 30]],
    ),
}


def _cell(value):
    """Excel gives datetimes and floats where the sheet showed plain text."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, (int, str)):
        return value
    return str(value)


def _sheet(name: str, rows: list[list]) -> dict:
    rows = [r for r in rows if any(c not in (None, "") for c in r)]
    if not rows:
        return {"name": name, "headers": [], "rows": []}
    width = max(len(r) for r in rows)
    padded = [list(r) + [None] * (width - len(r)) for r in rows]
    headers = [
        str(h).strip() if h not in (None, "") else f"Column {get_column_letter(i + 1)}"
        for i, h in enumerate(padded[0])
    ]
    return {"name": name, "headers": headers, "rows": padded[1 : MAX_ROWS + 1]}


def parse_upload(filename: str, blob: bytes) -> dict:
    """Return every sheet as headers + rows. Mapping happens in the browser."""
    if len(blob) > MAX_BYTES:
        raise ValueError(f"File is larger than {MAX_BYTES // (1024 * 1024)} MB")

    if filename.lower().endswith((".csv", ".txt")):
        text = blob.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = [[_cell(c) for c in r] for r in csv.reader(io.StringIO(text), dialect)]
        return {"sheets": [_sheet(filename.rsplit(".", 1)[0], rows)]}

    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ValueError("Upload a .xlsx or .csv file")

    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        sheets = []
        for ws in wb.worksheets:
            rows = [
                [_cell(c) for c in row]
                for row in ws.iter_rows(max_row=MAX_ROWS + 1, values_only=True)
            ]
            sheets.append(_sheet(ws.title, rows))
    finally:
        wb.close()
    if not sheets:
        raise ValueError("That workbook has no sheets")
    return {"sheets": sheets}


def _style_header(ws, width: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    for i in range(1, width + 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        ws.column_dimensions[get_column_letter(i)].width = max(14, len(str(c.value)) + 4)
    ws.freeze_panes = "A2"


def template_bytes() -> bytes:
    """A workbook shaped exactly like what the importer expects, with examples."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, (headers, examples) in TEMPLATE.items():
        ws = wb.create_sheet(title)
        ws.append(headers)
        for row in examples:
            ws.append(row)
        _style_header(ws, len(headers))

    notes = wb.create_sheet("Read me")
    for line in [
        ["EduSchedule import template"],
        [],
        ["Replace the example rows with your own. Delete rows you do not need."],
        ["Sheet and column names do not have to match — you map them after uploading."],
        [],
        ["can_teach", "Subject codes separated by ; or ,  (they must exist on the Subjects sheet)"],
        ["unavailable", "Slots the teacher cannot teach, e.g. 'Mon P1; Fri P6'. Blank means always free."],
        ["block_size", "Consecutive periods per session. 3 for a KTU lab. hours_per_week must divide by it."],
        ["room_type", "Free text, but it must exactly match the type of at least one room."],
        ["is_heavy", "yes / no. Heavy subjects get spread apart where the solver can manage it."],
    ]:
        notes.append(line)
    notes.column_dimensions["A"].width = 16
    notes.column_dimensions["B"].width = 90
    notes["A1"].font = Font(bold=True, size=14)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def solution_bytes(solution: dict) -> bytes:
    """Flat Class/Day/Period/Subject/Teacher/Room sheet, plus one grid per class."""
    wb = Workbook()
    ws = wb.active
    ws.title = "All classes"
    ws.append(["Class", "Day", "Period", "Subject", "Teacher", "Room"])
    for cls, table in solution.get("class_timetables", {}).items():
        for d, day in enumerate(table):
            for slot in day:
                ws.append([
                    cls,
                    DAY_NAMES[d] if d < len(DAY_NAMES) else f"Day {d + 1}",
                    f"P{slot['period']}",
                    slot["subject"] or "Free",
                    slot["teacher"] or "-",
                    slot["room"] or "-",
                ])
    _style_header(ws, 6)

    for cls, table in solution.get("class_timetables", {}).items():
        # sheet titles cannot contain : \ / ? * [ ] and cap at 31 chars
        safe = "".join(ch for ch in cls if ch not in ':\\/?*[]')[:31]
        grid = wb.create_sheet(safe or "Class")
        periods = len(table[0]) if table else 0
        grid.append([""] + [f"P{p + 1}" for p in range(periods)])
        for d, day in enumerate(table):
            grid.append(
                [DAY_NAMES[d] if d < len(DAY_NAMES) else f"Day {d + 1}"]
                + [
                    f"{s['subject']}\n{s['teacher'] or ''} · {s['room'] or ''}" if s["subject"] else ""
                    for s in day
                ]
            )
        _style_header(grid, periods + 1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
